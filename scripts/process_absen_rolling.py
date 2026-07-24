"""
Skrip Sinkronisasi Absen Excel ke LMS
-------------------------------------
Input: D:\\belajar\\5. ABSEN XII_XI_X Ganjil 26-27 OK.xlsx

Proses:
1. Kelas X: Dibuatkan file import Excel (karena belum punya akun)
2. Kelas XI & XII: Dibuatkan skrip sinkronisasi PHP untuk langsung update (pindah kelas) di database LMS
"""

import os
import json
import openpyxl

def process_absen_and_sync(input_path, output_dir, tp="2026/2027"):
    if not os.path.exists(input_path):
        print(f"Error: File '{input_path}' tidak ditemukan.")
        return False

    print(f"Memuat file absen: {input_path} ...")
    wb = openpyxl.load_workbook(input_path, data_only=True)
    class_sheets = [s for s in wb.sheetnames if s.startswith(('X', 'XI', 'XII'))]

    kelas_x_students = []
    kelas_xi_xii_students = []

    for sheet_name in class_sheets:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        for r in rows[9:]: # row 10 onwards
            if r and len(r) > 5 and r[4] and r[5]:
                nisn = str(r[2]).strip() if r[2] is not None else ''
                nis = str(r[3]).strip() if r[3] is not None else ''
                name = str(r[4]).strip() if r[4] is not None else ''
                jk = str(r[5]).strip().upper() if r[5] is not None else ''
                
                if name and name.upper() not in ('NAMA', 'NAMA S') and name.upper() != 'NONE':
                    lms_class_name = sheet_name.replace('_', '.').replace('-', '.')
                    student_data = {
                        'nisn': nisn,
                        'nis': nis,
                        'name': name,
                        'jk': jk,
                        'class_name': lms_class_name,
                        'academic_year': tp
                    }
                    
                    if sheet_name.startswith('X') and not sheet_name.startswith('XI'):
                        kelas_x_students.append(student_data)
                    else:
                        kelas_xi_xii_students.append(student_data)

    print(f"Ditemukan: {len(kelas_x_students)} Siswa Kelas X (Siswa Baru)")
    print(f"Ditemukan: {len(kelas_xi_xii_students)} Siswa Kelas XI & XII (Siswa Lama)\n")

    # 1. BUAT FILE IMPORT UNTUK KELAS X SAJA (Akun Baru)
    import_path_x = os.path.join(output_dir, f"import_siswa_lms_KELAS_X_{tp.replace('/', '_')}.xlsx")
    generate_lms_import_excel(kelas_x_students, import_path_x, tp, "Siswa Baru Kelas X")

    # 2. BUAT FILE IMPORT UNTUK KELAS XI & XII SAJA (Update Pindah Kelas)
    import_path_xi_xii = os.path.join(output_dir, f"import_siswa_lms_KELAS_XI_XII_{tp.replace('/', '_')}.xlsx")
    generate_lms_import_excel(kelas_xi_xii_students, import_path_xi_xii, tp, "Siswa Pindah Kelas XI XII")
    
    print("\n[OK] File Import telah sukses dipisahkan menjadi dua file siap pakai untuk LMS.")
    return True

def generate_lms_import_excel(students, output_path, tp, sheet_title="Siswa"):
    wb_lms = openpyxl.Workbook()
    ws = wb_lms.active
    ws.title = sheet_title
    
    headers = ["Nama", "Email", "Password", "Role", "NISN", "NIS", "Jenis Kelamin", "Kelas", "Tahun Ajaran"]
    ws.append(headers)

    count = 0
    for st in students:
        safe_id = st['nisn'] or st['nis'] or f"siswa_{count+1}"
        email = f"{safe_id}@siswa.belajar.id"
        default_password = f"Siswa{safe_id[-4:] if len(safe_id)>=4 else '1234'}"
        
        ws.append([
            st['name'], email, default_password, 'siswa',
            st['nisn'], st['nis'], st['jk'],
            st['class_name'], tp
        ])
        count += 1

    wb_lms.save(output_path)
    print(f"[OK] File Import ({count} siswa) dibuat di: {output_path}")

if __name__ == '__main__':
    input_file = r"D:\belajar\5. ABSEN XII_XI_X Ganjil 26-27 OK.xlsx"
    output_dir = r"D:\belajar"
    process_absen_and_sync(input_file, output_dir, tp="2026/2027")
